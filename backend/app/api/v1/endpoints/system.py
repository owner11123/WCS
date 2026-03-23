from typing import List, Any
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import FileResponse, StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import text
import os
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

from app.db.session import SessionLocal
from app.models.user import User

# Dependency
def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()
from app.schemas.user import User as UserSchema, UserCreate, UserUpdate
from app.core.security import get_password_hash, verify_password
from pydantic import BaseModel

router = APIRouter()

def get_app_dir():
    return os.path.abspath(os.path.join(os.path.dirname(__file__), "../../../.."))

def get_template_dir():
    return os.path.join(get_app_dir(), "templates")

def get_outbound_request_template_path():
    return os.path.join(get_template_dir(), "outbound_request_template.xlsx")

# --- User Management ---

@router.get("/users", response_model=List[UserSchema])
def read_users(skip: int = 0, limit: int = 100, db: Session = Depends(get_db)):
    users = db.query(User).offset(skip).limit(limit).all()
    return users

@router.post("/users", response_model=UserSchema)
def create_user(user_in: UserCreate, db: Session = Depends(get_db)):
    user = db.query(User).filter(User.username == user_in.username).first()
    if user:
        raise HTTPException(status_code=400, detail="Username already registered")
    
    db_user = User(
        username=user_in.username,
        password_hash=get_password_hash(user_in.password),
        role=user_in.role,
        is_active=user_in.is_active
    )
    db.add(db_user)
    db.commit()
    db.refresh(db_user)
    return db_user

@router.put("/users/{user_id}", response_model=UserSchema)
def update_user(user_id: int, user_in: UserUpdate, db: Session = Depends(get_db)):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    user.username = user_in.username
    user.role = user_in.role
    user.is_active = user_in.is_active
    if user_in.password:
        user.password_hash = get_password_hash(user_in.password)
        
    db.commit()
    db.refresh(user)
    return user

@router.delete("/users/{user_id}")
def delete_user(user_id: int, db: Session = Depends(get_db)):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Soft delete or hard delete depending on policy, we do hard delete for now
    db.delete(user)
    db.commit()
    return {"message": "User deleted successfully"}

# --- Data Management ---

class ClearDataRequest(BaseModel):
    password: str

@router.post("/clear-data")
def clear_system_data(req: ClearDataRequest, db: Session = Depends(get_db)):
    # 1. Verify admin password
    admin = db.query(User).filter(User.username == "admin").first()
    if not admin or not verify_password(req.password, admin.password_hash):
        raise HTTPException(status_code=403, detail="Invalid admin password")

    try:
        # 2. Truncate business data tables
        # Assuming PostgreSQL, CASCADE is needed to bypass foreign key constraints
        db.execute(text("TRUNCATE TABLE stock_transaction CASCADE;"))
        db.execute(text("TRUNCATE TABLE outbound_order CASCADE;"))
        db.execute(text("TRUNCATE TABLE inbound_order CASCADE;"))
        db.execute(text("TRUNCATE TABLE stock CASCADE;"))
        db.execute(text("TRUNCATE TABLE material_price_version CASCADE;"))
        db.execute(text("TRUNCATE TABLE location CASCADE;"))
        db.execute(text("TRUNCATE TABLE material CASCADE;"))
        
        db.commit()
        return {"message": "System data cleared successfully"}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

# --- Template Management ---

@router.get("/templates/outbound-request")
def download_outbound_request_template():
    path = get_outbound_request_template_path()
    if not os.path.exists(path):
        wb = Workbook()
        ws = wb.active
        ws.title = "领用申请单"

        ws.page_setup.paperSize = 9
        ws.page_setup.orientation = "portrait"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.35, bottom=0.35, header=0.2, footer=0.2)

        col_widths = [6, 14, 26, 18, 8, 8, 12, 14, 18]
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        thick = Side(style="medium", color="000000")
        border_thick = Border(left=thick, right=thick, top=thick, bottom=thick)

        ws.merge_cells("A1:I1")
        ws["A1"].value = "赢联盟西芒杜矿山项目公司寄售件物资领用申请单"
        ws["A1"].font = Font(size=18, bold=True)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        ws.merge_cells("F2:G2")
        ws.merge_cells("H2:I2")
        ws["F2"].value = "单据编号："
        ws["H2"].value = ""
        ws["F2"].alignment = Alignment(horizontal="right", vertical="center")
        ws["H2"].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[2].height = 20

        ws.merge_cells("A3:E3")
        ws.merge_cells("F3:I3")
        ws["A3"].value = "申领部门/单位："
        ws["F3"].value = "领用日期："
        ws["A3"].alignment = Alignment(horizontal="left", vertical="center")
        ws["F3"].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[3].height = 22

        headers_row = 4
        headers = ["序号", "物料编码", "物资名称", "型号/备件号", "品牌", "单位", "申请数量", "使用设备", "备注"]
        for c, name in enumerate(headers, start=1):
            cell = ws.cell(row=headers_row, column=c, value=name)
            cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border_thick
        ws.row_dimensions[headers_row].height = 22

        start_row = 5
        rows = 10
        for i in range(rows):
            r = start_row + i
            for c in range(1, 10):
                cell = ws.cell(row=r, column=c, value="")
                cell.border = border_thick
                cell.alignment = Alignment(horizontal="center" if c in (1, 5, 6, 7, 8) else "left", vertical="center", wrap_text=True)
                cell.font = Font(size=12)
            ws.row_dimensions[r].height = 26

        footer_row = start_row + rows
        ws.merge_cells(f"A{footer_row}:C{footer_row}")
        ws.merge_cells(f"D{footer_row}:F{footer_row}")
        ws.merge_cells(f"G{footer_row}:I{footer_row}")
        ws[f"A{footer_row}"].value = "领用人："
        ws[f"D{footer_row}"].value = "仓储复核："
        ws[f"G{footer_row}"].value = "发货人："
        for c in range(1, 10):
            cell = ws.cell(row=footer_row, column=c)
            cell.border = border_thick
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.font = Font(size=12)
        ws.row_dimensions[footer_row].height = 24

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        headers = {"Content-Disposition": "attachment; filename=\"outbound_request_template.xlsx\""}
        return StreamingResponse(
            iter([output.getvalue()]),
            headers=headers,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    return FileResponse(
        path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="outbound_request_template.xlsx"
    )

@router.post("/templates/outbound-request")
async def upload_outbound_request_template(file: UploadFile = File(...)):
    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx is allowed")
    content = await file.read()
    template_dir = get_template_dir()
    os.makedirs(template_dir, exist_ok=True)
    path = get_outbound_request_template_path()
    with open(path, "wb") as f:
        f.write(content)
    return {"message": "Template uploaded successfully"}
