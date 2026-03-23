from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
import pandas as pd
import io
from io import BytesIO
from datetime import datetime
from urllib.parse import quote
import os
from datetime import timedelta
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

from app.db.session import SessionLocal
from app.models.material import Material, MaterialPriceVersion
from app.models.transit import TransitInventory
from app.models.order import InboundOrder, OutboundOrder
from app.models.location import Location
from app.models.stock import Stock, StockTransaction
from app.models.user import User

router = APIRouter()

def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

def get_app_dir():
    return os.path.abspath(os.path.join(os.path.dirname(__file__), "../../../.."))

def get_template_path():
    return os.path.join(get_app_dir(), "templates", "outbound_request_template.xlsx")

def parse_start_end(start_date: str, end_date: str):
    start_dt = datetime.strptime(start_date, "%Y-%m-%d") if start_date else None
    end_dt = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1) if end_date else None
    return start_dt, end_dt

@router.post("/import/materials")
async def import_materials(file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Invalid file format. Please upload an Excel file.")
    
    try:
        contents = await file.read()
        df = pd.read_excel(BytesIO(contents))
        
        # Expected columns: code, model, description, category_major, category_minor, substitute_code, vehicle_model
        for index, row in df.iterrows():
            if pd.isna(row.get('code')):
                continue
                
            code = str(row.get('code', '')).strip()
            if not code or code == 'nan':
                continue
                
            # Check if material exists
            existing = db.query(Material).filter(Material.code == code).first()
            
            substitute_code = str(row.get('substitute_code', '')).strip()
            if substitute_code == 'nan': substitute_code = ''
            category_major = str(row.get('category_major', '')).strip().replace('nan', '')
            category_minor = str(row.get('category_minor', '')).strip().replace('nan', '')
            if category_major == 'nan': category_major = ''
            if category_minor == 'nan': category_minor = ''

            if not existing:
                material = Material(
                    code=code,
                    model=str(row.get('model', '')).replace('nan', ''),
                    description=str(row.get('description', '')).replace('nan', ''),
                    category_major=category_major,
                    category_minor=category_minor,
                    substitute_code=substitute_code,
                    vehicle_model=str(row.get('vehicle_model', '')).replace('nan', '')
                )
                db.add(material)
            else:
                existing.model = str(row.get('model', '')).replace('nan', '')
                existing.description = str(row.get('description', '')).replace('nan', '')
                existing.category_major = category_major
                existing.category_minor = category_minor
                existing.substitute_code = substitute_code
                existing.vehicle_model = str(row.get('vehicle_model', '')).replace('nan', '')
        
        db.commit()
        return {"message": "Materials imported successfully"}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/template/transit")
def template_transit():
    columns = ["箱号", "物料编码", "物料描述", "适用车型", "采购合同号", "数量", "采购单价", "销售单价", "货币"]
    df = pd.DataFrame(columns=columns)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='在途库存导入模板')
    output.seek(0)
    headers = {
        'Content-Disposition': 'attachment; filename="transit_template.xlsx"'
    }
    return StreamingResponse(
        iter([output.getvalue()]),
        headers=headers,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@router.get("/print/outbound/{group_no}")
def print_outbound_request_excel(group_no: str, rows: int = 10, db: Session = Depends(get_db)):
    try:
        records = db.query(
            OutboundOrder.group_no,
            OutboundOrder.customer,
            OutboundOrder.receiver,
            OutboundOrder.outbound_time,
            User.username.label("operator_name"),
            Material.code.label("material_code"),
            Material.model.label("material_model"),
            Material.description.label("material_description"),
            Material.vehicle_model.label("vehicle_model"),
            MaterialPriceVersion.batch_no.label("contract_no"),
            OutboundOrder.quantity
        ).join(
            Material, OutboundOrder.material_id == Material.id
        ).outerjoin(
            User, OutboundOrder.operator_id == User.id
        ).join(
            MaterialPriceVersion, OutboundOrder.price_version_id == MaterialPriceVersion.id
        ).filter(
            OutboundOrder.group_no == group_no
        ).order_by(OutboundOrder.id.asc()).all()

        if not records:
            raise HTTPException(status_code=404, detail="Outbound order not found")

        doc_no = records[0].group_no
        customer = records[0].customer or ""
        outbound_time = records[0].outbound_time or datetime.utcnow()
        operator_name = records[0].operator_name or ""
        date_str = outbound_time.strftime("%Y.%m.%d")

        default_brand = "徐工"
        default_unit = "个"

        template_path = get_template_path()
        if os.path.exists(template_path):
            wb = load_workbook(template_path)
            ws = wb.active

            ws["H2"].value = doc_no
            if not ws["A3"].value or str(ws["A3"].value).startswith("申领部门/单位"):
                ws["A3"].value = f"申领部门/单位：{customer}"
            if not ws["F3"].value or str(ws["F3"].value).startswith("领用日期"):
                ws["F3"].value = f"领用日期：{date_str}"

            data_start_row = 5
            footer_row_base = None
            for r in range(1, ws.max_row + 1):
                v = ws.cell(row=r, column=1).value
                if v and "领用人" in str(v):
                    footer_row_base = r
                    break
            if footer_row_base:
                max_rows_in_template = max(footer_row_base - data_start_row, 1)
            else:
                max_rows_in_template = max(int(rows or 0), 10)

            target_rows = max(max_rows_in_template, len(records), 1)
            if footer_row_base and target_rows > max_rows_in_template:
                insert_n = target_rows - max_rows_in_template
                ws.insert_rows(footer_row_base, insert_n)
                footer_row = footer_row_base + insert_n
            else:
                footer_row = footer_row_base or (data_start_row + target_rows)

            for i in range(target_rows):
                r = data_start_row + i
                rec = records[i] if i < len(records) else None
                ws.cell(row=r, column=1).value = i + 1 if rec else ""
                ws.cell(row=r, column=2).value = (rec.material_model if rec else "") or ""
                ws.cell(row=r, column=3).value = (rec.material_description if rec else "") or ""
                ws.cell(row=r, column=4).value = (rec.material_code if rec else "") or ""
                if rec:
                    if not ws.cell(row=r, column=5).value:
                        ws.cell(row=r, column=5).value = default_brand
                    if not ws.cell(row=r, column=6).value:
                        ws.cell(row=r, column=6).value = default_unit
                else:
                    ws.cell(row=r, column=5).value = ws.cell(row=r, column=5).value or ""
                    ws.cell(row=r, column=6).value = ws.cell(row=r, column=6).value or ""
                ws.cell(row=r, column=7).value = int(rec.quantity) if rec else ""
                ws.cell(row=r, column=8).value = (rec.vehicle_model if rec else "") or ""
                ws.cell(row=r, column=9).value = (rec.contract_no if rec else "") or ""

            if footer_row:
                for c in range(1, 10):
                    cell = ws.cell(row=footer_row, column=c)
                    if cell.value and "发货人" in str(cell.value):
                        cell.value = f"发货人：{operator_name} {date_str}"
                        break
                else:
                    ws.cell(row=footer_row, column=7).value = f"发货人：{operator_name} {date_str}"
        else:
            target_rows = max(int(rows or 0), len(records), 1)
            wb = Workbook()
            ws = wb.active
            ws.title = "领用申请单"

            ws.page_setup.paperSize = 9
            ws.page_setup.orientation = "portrait"
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0
            ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.35, bottom=0.35, header=0.2, footer=0.2)

            col_widths = [6, 14, 26, 18, 8, 8, 12, 14, 18]
            for i, w in enumerate(col_widths, start=1):
                ws.column_dimensions[get_column_letter(i)].width = w

            thick = Side(style="medium", color="000000")
            border_thick = Border(left=thick, right=thick, top=thick, bottom=thick)

            ws.merge_cells("A1:I1")
            ws["A1"].value = "赢联盟西芒杜矿山项目公司寄售件物资领用申请单"
            ws["A1"].font = Font(size=18, bold=True)
            ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 30

            ws.merge_cells("F2:G2")
            ws.merge_cells("H2:I2")
            ws["F2"].value = "单据编号："
            ws["H2"].value = doc_no
            ws["F2"].alignment = Alignment(horizontal="right", vertical="center")
            ws["H2"].alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[2].height = 20

            ws.merge_cells("A3:E3")
            ws.merge_cells("F3:I3")
            ws["A3"].value = f"申领部门/单位：{customer}"
            ws["F3"].value = f"领用日期：{date_str}"
            ws["A3"].alignment = Alignment(horizontal="left", vertical="center")
            ws["F3"].alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[3].height = 22

            headers_row = 4
            headers = ["序号", "物料编码", "物资名称", "型号/备件号", "品牌", "单位", "申请数量", "使用设备", "备注"]
            for c, name in enumerate(headers, start=1):
                cell = ws.cell(row=headers_row, column=c, value=name)
                cell.font = Font(bold=True, size=12)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border_thick
            ws.row_dimensions[headers_row].height = 22

            start_row = 5
            for i in range(target_rows):
                r = start_row + i
                rec = records[i] if i < len(records) else None
                values = [
                    i + 1 if rec else "",
                    (rec.material_model if rec else "") or "",
                    (rec.material_description if rec else "") or "",
                    (rec.material_code if rec else "") or "",
                    default_brand if rec else "",
                    default_unit if rec else "",
                    int(rec.quantity) if rec else "",
                    (rec.vehicle_model if rec else "") or "",
                    (rec.contract_no if rec else "") or ""
                ]
                for c, val in enumerate(values, start=1):
                    cell = ws.cell(row=r, column=c, value=val)
                    cell.border = border_thick
                    if c in (1, 5, 6, 7, 8):
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                    cell.font = Font(size=12)
                ws.row_dimensions[r].height = 26

            footer_row = start_row + target_rows
            ws.merge_cells(f"A{footer_row}:C{footer_row}")
            ws.merge_cells(f"D{footer_row}:F{footer_row}")
            ws.merge_cells(f"G{footer_row}:I{footer_row}")
            ws[f"A{footer_row}"].value = "领用人："
            ws[f"D{footer_row}"].value = "仓储复核："
            ws[f"G{footer_row}"].value = f"发货人：{operator_name} {date_str}"
            for c in range(1, 10):
                cell = ws.cell(row=footer_row, column=c)
                cell.border = border_thick
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.font = Font(size=12)
            ws.row_dimensions[footer_row].height = 24

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"{doc_no}_领用申请单.xlsx"
        headers = {
            "Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"
        }
        return StreamingResponse(
            iter([output.getvalue()]),
            headers=headers,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/export/transactions")
def export_transactions(start_date: str = None, end_date: str = None, db: Session = Depends(get_db)):
    start_dt, end_dt = parse_start_end(start_date, end_date)

    query = db.query(
        StockTransaction.created_at.label("transaction_time"),
        StockTransaction.transaction_type.label("transaction_type"),
        Material.code.label("material_code"),
        Material.description.label("material_description"),
        MaterialPriceVersion.batch_no.label("batch_no"),
        Location.code.label("location_code"),
        StockTransaction.quantity_change.label("quantity_change"),
        StockTransaction.balance.label("balance"),
        StockTransaction.reference_order.label("reference_order"),
        User.username.label("operator_name")
    ).join(
        Material, StockTransaction.material_id == Material.id
    ).join(
        Location, StockTransaction.location_id == Location.id
    ).outerjoin(
        MaterialPriceVersion, StockTransaction.price_version_id == MaterialPriceVersion.id
    ).outerjoin(
        User, StockTransaction.operator_id == User.id
    )

    if start_dt:
        query = query.filter(StockTransaction.created_at >= start_dt)
    if end_dt:
        query = query.filter(StockTransaction.created_at < end_dt)

    records = query.order_by(StockTransaction.created_at.desc()).all()
    data = []
    for r in records:
        data.append({
            "操作时间": r.transaction_time.strftime("%Y-%m-%d %H:%M:%S") if r.transaction_time else "",
            "操作类型": r.transaction_type,
            "物料编码": r.material_code,
            "物料描述": r.material_description,
            "批次/合同号": r.batch_no,
            "库位": r.location_code,
            "变动数量": r.quantity_change,
            "结余库存": r.balance,
            "关联单号": r.reference_order,
            "操作人": r.operator_name
        })

    df = pd.DataFrame(data)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='操作流水')
    output.seek(0)

    name_start = start_date or ""
    name_end = end_date or ""
    filename = f"transactions_{name_start}_{name_end}.xlsx" if (name_start or name_end) else "transactions.xlsx"
    headers = {"Content-Disposition": f"attachment; filename=\"{filename}\""}
    return StreamingResponse(
        iter([output.getvalue()]),
        headers=headers,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@router.get("/print/outbound-batch")
def print_outbound_request_excel_batch(customer: str, start_date: str = None, end_date: str = None, rows: int = 10, db: Session = Depends(get_db)):
    start_dt, end_dt = parse_start_end(start_date, end_date)

    q = db.query(OutboundOrder.group_no).filter(OutboundOrder.group_no.isnot(None))
    q = q.filter(OutboundOrder.customer.ilike(f"%{customer}%"))
    if start_dt:
        q = q.filter(OutboundOrder.outbound_time >= start_dt)
    if end_dt:
        q = q.filter(OutboundOrder.outbound_time < end_dt)

    group_nos = [r[0] for r in q.distinct().order_by(OutboundOrder.group_no.asc()).all() if r[0]]
    if not group_nos:
        raise HTTPException(status_code=404, detail="No outbound orders found for criteria")

    default_brand = "徐工"
    default_unit = "个"
    template_path = get_template_path()

    def fetch_records(gno: str):
        return db.query(
            OutboundOrder.group_no,
            OutboundOrder.customer,
            OutboundOrder.receiver,
            OutboundOrder.outbound_time,
            User.username.label("operator_name"),
            Material.code.label("material_code"),
            Material.model.label("material_model"),
            Material.description.label("material_description"),
            Material.vehicle_model.label("vehicle_model"),
            MaterialPriceVersion.batch_no.label("contract_no"),
            OutboundOrder.quantity
        ).join(
            Material, OutboundOrder.material_id == Material.id
        ).outerjoin(
            User, OutboundOrder.operator_id == User.id
        ).join(
            MaterialPriceVersion, OutboundOrder.price_version_id == MaterialPriceVersion.id
        ).filter(
            OutboundOrder.group_no == gno
        ).order_by(OutboundOrder.id.asc()).all()

    def fill_sheet(ws, records):
        doc_no = records[0].group_no
        customer_val = records[0].customer or ""
        outbound_time = records[0].outbound_time or datetime.utcnow()
        operator_name = records[0].operator_name or ""
        date_str = outbound_time.strftime("%Y.%m.%d")

        ws["H2"].value = doc_no
        if ws["A3"].value is None or str(ws["A3"].value).startswith("申领部门/单位"):
            ws["A3"].value = f"申领部门/单位：{customer_val}"
        if ws["F3"].value is None or str(ws["F3"].value).startswith("领用日期"):
            ws["F3"].value = f"领用日期：{date_str}"

        data_start_row = 5
        footer_row_base = None
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=1).value
            if v and "领用人" in str(v):
                footer_row_base = r
                break
        if footer_row_base:
            max_rows_in_template = max(footer_row_base - data_start_row, 1)
        else:
            max_rows_in_template = max(int(rows or 0), 10)

        target_rows = max(max_rows_in_template, len(records), max(int(rows or 0), 1))
        if footer_row_base and target_rows > max_rows_in_template:
            insert_n = target_rows - max_rows_in_template
            ws.insert_rows(footer_row_base, insert_n)
            footer_row = footer_row_base + insert_n
        else:
            footer_row = footer_row_base or (data_start_row + target_rows)

        for i in range(target_rows):
            r = data_start_row + i
            rec = records[i] if i < len(records) else None
            ws.cell(row=r, column=1).value = i + 1 if rec else ""
            ws.cell(row=r, column=2).value = (rec.material_model if rec else "") or ""
            ws.cell(row=r, column=3).value = (rec.material_description if rec else "") or ""
            ws.cell(row=r, column=4).value = (rec.material_code if rec else "") or ""
            if rec:
                if not ws.cell(row=r, column=5).value:
                    ws.cell(row=r, column=5).value = default_brand
                if not ws.cell(row=r, column=6).value:
                    ws.cell(row=r, column=6).value = default_unit
            else:
                ws.cell(row=r, column=5).value = ws.cell(row=r, column=5).value or ""
                ws.cell(row=r, column=6).value = ws.cell(row=r, column=6).value or ""
            ws.cell(row=r, column=7).value = int(rec.quantity) if rec else ""
            ws.cell(row=r, column=8).value = (rec.vehicle_model if rec else "") or ""
            ws.cell(row=r, column=9).value = (rec.contract_no if rec else "") or ""

        if footer_row:
            updated = False
            for c in range(1, 10):
                cell = ws.cell(row=footer_row, column=c)
                if cell.value and "发货人" in str(cell.value):
                    cell.value = f"发货人：{operator_name} {date_str}"
                    updated = True
                    break
            if not updated:
                ws.cell(row=footer_row, column=7).value = f"发货人：{operator_name} {date_str}"

    if os.path.exists(template_path):
        wb = load_workbook(template_path)
        base_ws = wb.active
        base_ws.title = group_nos[0][:31]
        first_records = fetch_records(group_nos[0])
        if not first_records:
            raise HTTPException(status_code=404, detail="Outbound order not found")
        fill_sheet(base_ws, first_records)

        for gno in group_nos[1:]:
            records = fetch_records(gno)
            if not records:
                continue
            ws = wb.copy_worksheet(base_ws)
            ws.title = gno[:31]
            fill_sheet(ws, records)
    else:
        wb = Workbook()
        wb.remove(wb.active)
        thick = Side(style="medium", color="000000")
        border_thick = Border(left=thick, right=thick, top=thick, bottom=thick)

        def build_sheet(ws, records):
            doc_no = records[0].group_no
            customer_val = records[0].customer or ""
            outbound_time = records[0].outbound_time or datetime.utcnow()
            operator_name = records[0].operator_name or ""
            date_str = outbound_time.strftime("%Y.%m.%d")
            target_rows = max(int(rows or 0), len(records), 1)

            ws.page_setup.paperSize = 9
            ws.page_setup.orientation = "portrait"
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0
            ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.35, bottom=0.35, header=0.2, footer=0.2)

            col_widths = [6, 14, 26, 18, 8, 8, 12, 14, 18]
            for i, w in enumerate(col_widths, start=1):
                ws.column_dimensions[get_column_letter(i)].width = w

            ws.merge_cells("A1:I1")
            ws["A1"].value = "赢联盟西芒杜矿山项目公司寄售件物资领用申请单"
            ws["A1"].font = Font(size=18, bold=True)
            ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 30

            ws.merge_cells("F2:G2")
            ws.merge_cells("H2:I2")
            ws["F2"].value = "单据编号："
            ws["H2"].value = doc_no
            ws["F2"].alignment = Alignment(horizontal="right", vertical="center")
            ws["H2"].alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[2].height = 20

            ws.merge_cells("A3:E3")
            ws.merge_cells("F3:I3")
            ws["A3"].value = f"申领部门/单位：{customer_val}"
            ws["F3"].value = f"领用日期：{date_str}"
            ws["A3"].alignment = Alignment(horizontal="left", vertical="center")
            ws["F3"].alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[3].height = 22

            headers_row = 4
            headers = ["序号", "物料编码", "物资名称", "型号/备件号", "品牌", "单位", "申请数量", "使用设备", "备注"]
            for c, name in enumerate(headers, start=1):
                cell = ws.cell(row=headers_row, column=c, value=name)
                cell.font = Font(bold=True, size=12)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border_thick
            ws.row_dimensions[headers_row].height = 22

            start_row = 5
            for i in range(target_rows):
                r = start_row + i
                rec = records[i] if i < len(records) else None
                values = [
                    i + 1 if rec else "",
                    (rec.material_model if rec else "") or "",
                    (rec.material_description if rec else "") or "",
                    (rec.material_code if rec else "") or "",
                    default_brand if rec else "",
                    default_unit if rec else "",
                    int(rec.quantity) if rec else "",
                    (rec.vehicle_model if rec else "") or "",
                    (rec.contract_no if rec else "") or ""
                ]
                for c, val in enumerate(values, start=1):
                    cell = ws.cell(row=r, column=c, value=val)
                    cell.border = border_thick
                    if c in (1, 5, 6, 7, 8):
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                    cell.font = Font(size=12)
                ws.row_dimensions[r].height = 26

            footer_row = start_row + target_rows
            ws.merge_cells(f"A{footer_row}:C{footer_row}")
            ws.merge_cells(f"D{footer_row}:F{footer_row}")
            ws.merge_cells(f"G{footer_row}:I{footer_row}")
            ws[f"A{footer_row}"].value = "领用人："
            ws[f"D{footer_row}"].value = "仓储复核："
            ws[f"G{footer_row}"].value = f"发货人：{operator_name} {date_str}"
            for c in range(1, 10):
                cell = ws.cell(row=footer_row, column=c)
                cell.border = border_thick
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.font = Font(size=12)
            ws.row_dimensions[footer_row].height = 24

        for gno in group_nos:
            records = fetch_records(gno)
            if not records:
                continue
            ws = wb.create_sheet(title=gno[:31])
            build_sheet(ws, records)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    name_start = start_date or ""
    name_end = end_date or ""
    filename = f"outbound_request_{customer}_{name_start}_{name_end}.xlsx" if (name_start or name_end) else f"outbound_request_{customer}.xlsx"
    headers = {"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"}
    return StreamingResponse(
        iter([output.getvalue()]),
        headers=headers,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@router.get("/template/inbound")
def template_inbound():
    columns = ["入库单号", "采购合同号", "物料编码", "物料型号", "物料描述", "适用车型", "库位编码", "入库数量", "采购单价", "销售单价", "货币", "入库时间"]
    df = pd.DataFrame(columns=columns)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='入库导入模板')
    output.seek(0)
    headers = {
        'Content-Disposition': 'attachment; filename="inbound_template.xlsx"'
    }
    return StreamingResponse(
        iter([output.getvalue()]),
        headers=headers,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@router.post("/import/transit")
async def import_transit(file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only Excel files are allowed")
    
    try:
        contents = await file.read()
        df = pd.read_excel(io.BytesIO(contents))
        
        imported_count = 0
        for index, row in df.iterrows():
            if pd.isna(row.get('箱号')) or pd.isna(row.get('物料编码')):
                continue
                
            box_no = str(row.get('箱号', '')).strip()
            mat_code = str(row.get('物料编码', '')).strip()
            qty = row.get('数量', 0)
            
            if not box_no or not mat_code or pd.isna(qty) or int(qty) <= 0:
                continue
                
            # 1. Update or Create Material
            material = db.query(Material).filter(Material.code == mat_code).first()
            mat_desc = str(row.get('物料描述', '')).strip() if not pd.isna(row.get('物料描述')) else None
            mat_vehicle = str(row.get('适用车型', '')).strip() if not pd.isna(row.get('适用车型')) else None
            
            if material:
                if mat_desc and mat_desc != 'nan': material.description = mat_desc
                if mat_vehicle and mat_vehicle != 'nan': material.vehicle_model = mat_vehicle
            else:
                material = Material(
                    code=mat_code,
                    description=mat_desc if mat_desc != 'nan' else None,
                    vehicle_model=mat_vehicle if mat_vehicle != 'nan' else None
                )
                db.add(material)
                db.flush()
                
            # 2. Update or Create Transit Inventory
            contract_no = str(row.get('采购合同号', '')).strip().replace('nan', '')
            if not contract_no: contract_no = "DEFAULT"
            
            purchase_price = row.get('采购单价')
            sale_price = row.get('销售单价')
            currency = str(row.get('货币', 'CNY')).strip().replace('nan', 'CNY')
            
            # Allow multiple materials per box_no by checking box_no + material_id combo
            transit_record = db.query(TransitInventory).filter(
                TransitInventory.box_no == box_no,
                TransitInventory.material_id == material.id,
                TransitInventory.status == "in_transit"
            ).first()
            
            if transit_record:
                # If same material exists in same box, add quantity
                add_qty = int(qty)
                transit_record.quantity += add_qty
                if transit_record.total_quantity is None:
                    base_total = (transit_record.quantity - add_qty) + (transit_record.received_quantity or 0)
                    transit_record.total_quantity = base_total + add_qty
                else:
                    transit_record.total_quantity += add_qty
                # Update other fields if provided
                if contract_no != "DEFAULT": transit_record.contract_no = contract_no
                if not pd.isna(purchase_price): transit_record.purchase_price = purchase_price
                if not pd.isna(sale_price): transit_record.sale_price = sale_price
            else:
                transit_record = TransitInventory(
                    box_no=box_no,
                    material_id=material.id,
                    contract_no=contract_no,
                    total_quantity=int(qty),
                    received_quantity=0,
                    quantity=int(qty),
                    purchase_price=purchase_price if not pd.isna(purchase_price) else None,
                    sale_price=sale_price if not pd.isna(sale_price) else None,
                    currency=currency,
                    status="in_transit"
                )
                db.add(transit_record)
            
            imported_count += 1
            
        db.commit()
        return {"message": f"Successfully imported {imported_count} transit inventory records"}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/import/inbound")
async def import_inbound(file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Invalid file format. Please upload an Excel file.")
    
    try:
        contents = await file.read()
        df = pd.read_excel(BytesIO(contents))
        
        base_order_no = f"IN-{datetime.utcnow().strftime('%Y%m%d%H%M%S')}"
        
        for index, row in df.iterrows():
            if pd.isna(row.get('物料编码')) or pd.isna(row.get('库位编码')):
                continue
                
            mat_code = str(row.get('物料编码', '')).strip()
            loc_code = str(row.get('库位编码', '')).strip()
            qty = row.get('入库数量', 0)
            
            # Handle inbound_time
            inbound_time_raw = row.get('入库时间')
            if pd.isna(inbound_time_raw) or not inbound_time_raw:
                inbound_time = datetime.utcnow()
            else:
                try:
                    inbound_time = pd.to_datetime(inbound_time_raw)
                except:
                    inbound_time = datetime.utcnow()
            
            if not mat_code or mat_code == 'nan' or not loc_code or loc_code == 'nan' or pd.isna(qty) or int(qty) <= 0:
                continue
                
            # 1. Check/Create Material
            material = db.query(Material).filter(Material.code == mat_code).first()
            if not material:
                material = Material(
                    code=mat_code,
                    model=str(row.get('物料型号', '')).replace('nan', ''),
                    description=str(row.get('物料描述', '')).replace('nan', ''),
                    vehicle_model=str(row.get('适用车型', '')).replace('nan', '')
                )
                db.add(material)
                db.flush()
            
            # 2. Check/Create Location
            location = db.query(Location).filter(Location.code == loc_code).first()
            if not location:
                location = Location(
                    code=loc_code
                )
                db.add(location)
                db.flush()
                
            # 3. Handle Price Version (Using contract_no as batch_no)
            contract_no = str(row.get('采购合同号', '')).strip().replace('nan', '')
            if not contract_no:
                contract_no = "DEFAULT"
                
            purchase_price = row.get('采购单价')
            sale_price = row.get('销售单价')
            currency = str(row.get('货币', 'CNY')).strip().replace('nan', 'CNY')
            
            if pd.isna(purchase_price): purchase_price = None
            if pd.isna(sale_price): sale_price = None
            
            # Find price version by material and contract_no (batch_no)
            pv = db.query(MaterialPriceVersion).filter(
                MaterialPriceVersion.material_id == material.id,
                MaterialPriceVersion.batch_no == contract_no
            ).first()
                
            if not pv:
                pv = MaterialPriceVersion(
                    material_id=material.id,
                    batch_no=contract_no,
                    purchase_price=purchase_price or 0,
                    sale_price=sale_price or 0,
                    currency=currency
                )
                db.add(pv)
                db.flush()
                
            # 4. Create Inbound Order
            order_no = str(row.get('入库单号', '')).strip().replace('nan', '')
            if not order_no:
                order_no = f"{base_order_no}-{index}"
                
            contract_no_value = contract_no if contract_no != "DEFAULT" else None
            
            order = InboundOrder(
                order_no=order_no,
                material_id=material.id,
                price_version_id=pv.id,
                location_id=location.id,
                quantity=int(qty),
                contract_no=contract_no_value,
                operator_id=1,
                status="completed",
                inbound_time=inbound_time
            )
            db.add(order)
            
            # 5. Update Stock
            stock = db.query(Stock).filter(
                Stock.material_id == material.id,
                Stock.location_id == location.id,
                Stock.price_version_id == pv.id
            ).first()
            
            if stock:
                stock.quantity += int(qty)
                stock.total_inbound += int(qty)
            else:
                stock = Stock(
                    material_id=material.id,
                    location_id=location.id,
                    price_version_id=pv.id,
                    quantity=int(qty),
                    total_inbound=int(qty),
                    total_outbound=0
                )
                db.add(stock)
            db.flush()
            
            # 6. Create Transaction Log
            transaction = StockTransaction(
                material_id=material.id,
                location_id=location.id,
                price_version_id=pv.id,
                transaction_type="inbound",
                quantity_change=int(qty),
                balance=stock.quantity,
                reference_order=order_no,
                operator_id=1
            )
            db.add(transaction)
            
        db.commit()
        return {"message": "Inbound records imported successfully"}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/template/outbound")
def template_outbound():
    columns = ["出库单号", "客户名称", "领用人", "合同号", "物料编码", "库位编码", "出库数量", "出库时间"]
    df = pd.DataFrame(columns=columns)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='出库导入模板')
    output.seek(0)
    headers = {
        'Content-Disposition': 'attachment; filename="outbound_template.xlsx"'
    }
    return StreamingResponse(
        iter([output.getvalue()]),
        headers=headers,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@router.post("/import/outbound")
async def import_outbound(file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Invalid file format. Please upload an Excel file.")
    
    try:
        contents = await file.read()
        df = pd.read_excel(BytesIO(contents))
        
        base_order_no = f"OUT-{datetime.utcnow().strftime('%Y%m%d%H%M%S')}"

        def clean_str(v):
            if v is None or (isinstance(v, float) and pd.isna(v)) or (hasattr(pd, "isna") and pd.isna(v)):
                return ""
            s = str(v).strip().replace("nan", "")
            if s.endswith(".0"):
                head = s[:-2]
                if head.isdigit():
                    return head
            s = s.rstrip(".").rstrip("。")
            return s

        def to_qty_int(v):
            if v is None or (isinstance(v, float) and pd.isna(v)) or (hasattr(pd, "isna") and pd.isna(v)):
                return 0
            s = str(v).replace(",", "").strip()
            if not s or s.lower() == "nan":
                return 0
            try:
                return int(float(s))
            except Exception:
                return 0
        
        for index, row in df.iterrows():
            if pd.isna(row.get('物料编码')) or pd.isna(row.get('库位编码')):
                continue
                
            mat_code = clean_str(row.get('物料编码', ''))
            loc_code = clean_str(row.get('库位编码', ''))
            qty = to_qty_int(row.get('出库数量', 0))

            outbound_time_raw = row.get('出库日期')
            if pd.isna(outbound_time_raw) or not outbound_time_raw:
                outbound_time_raw = row.get('出库时间')
            if pd.isna(outbound_time_raw) or not outbound_time_raw:
                outbound_time = datetime.utcnow()
            else:
                try:
                    outbound_time = pd.to_datetime(outbound_time_raw)
                except:
                    outbound_time = datetime.utcnow()
            
            if not mat_code or not loc_code or qty <= 0:
                continue
                
            material = db.query(Material).filter(Material.code == mat_code).first()
            if not material:
                raise HTTPException(status_code=400, detail=f"Material code {mat_code} not found.")
                
            location = db.query(Location).filter(Location.code == loc_code).first()
            if not location:
                raise HTTPException(status_code=400, detail=f"Location code {loc_code} not found.")

            contract_no = clean_str(row.get('合同号', ''))
            if not contract_no:
                contract_no = clean_str(row.get('采购合同号', ''))

            pv = None
            if contract_no:
                pv = db.query(MaterialPriceVersion).filter(
                    MaterialPriceVersion.material_id == material.id,
                    MaterialPriceVersion.batch_no == contract_no
                ).first()
                if not pv:
                    raise HTTPException(status_code=400, detail=f"Contract {contract_no} for material {mat_code} not found.")

            stock_query = db.query(Stock).filter(
                Stock.material_id == material.id,
                Stock.location_id == location.id,
                Stock.quantity >= int(qty)
            )
            if pv:
                stock_query = stock_query.filter(Stock.price_version_id == pv.id)
            stock = stock_query.with_for_update().first()
            
            if not stock:
                raise HTTPException(status_code=400, detail=f"Insufficient stock for material {mat_code} at location {loc_code}.")
                
            order_no = str(row.get('出库单号', '')).strip().replace('nan', '')
            if not order_no:
                order_no = f"{base_order_no}-{index}"
                
            customer = clean_str(row.get('客户名称', ''))
            receiver = clean_str(row.get('领用人', ''))
            
            order = OutboundOrder(
                order_no=order_no,
                material_id=material.id,
                price_version_id=stock.price_version_id,
                location_id=location.id,
                quantity=int(qty),
                customer=customer if customer else None,
                receiver=receiver if receiver else None,
                operator_id=1,
                status="completed",
                outbound_time=outbound_time
            )
            db.add(order)
            
            stock.quantity -= int(qty)
            stock.total_outbound += int(qty)
            db.flush()
            
            transaction = StockTransaction(
                material_id=material.id,
                location_id=location.id,
                price_version_id=stock.price_version_id,
                transaction_type="outbound",
                quantity_change=-int(qty),
                balance=stock.quantity,
                reference_order=order_no,
                operator_id=1
            )
            db.add(transaction)
            
        db.commit()
        return {"message": "Outbound records imported successfully"}
    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/export/materials")
def export_materials(db: Session = Depends(get_db)):
    materials = db.query(Material).filter(Material.is_deleted == False).all()
    
    data = []
    for m in materials:
        data.append({
            "Code": m.code,
            "Model": m.model,
            "Description": m.description,
            "Category Major": m.category_major,
            "Category Minor": m.category_minor,
            "Substitute Code": m.substitute_code,
            "Vehicle Model": m.vehicle_model
        })
        
    df = pd.DataFrame(data)
    
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Materials')
    
    output.seek(0)
    
    headers = {
        'Content-Disposition': 'attachment; filename="materials.xlsx"'
    }
    
    return StreamingResponse(
        iter([output.getvalue()]),
        headers=headers,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@router.get("/export/inbound")
def export_inbound(db: Session = Depends(get_db)):
    orders = db.query(InboundOrder, Material.code, Material.description, Location.code.label('loc_code')).join(
        Material, InboundOrder.material_id == Material.id
    ).join(
        Location, InboundOrder.location_id == Location.id
    ).all()
    
    data = []
    for order, mat_code, mat_desc, loc_code in orders:
        data.append({
            "入库单号": order.order_no,
            "物料编码": mat_code,
            "物料描述": mat_desc,
            "库位编码": loc_code,
            "入库数量": order.quantity,
            "状态": order.status,
            "入库时间": order.inbound_time.strftime("%Y-%m-%d %H:%M:%S") if order.inbound_time else ""
        })
        
    df = pd.DataFrame(data)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='入库记录')
    
    output.seek(0)
    headers = {
        'Content-Disposition': 'attachment; filename="inbound_records.xlsx"'
    }
    return StreamingResponse(
        iter([output.getvalue()]),
        headers=headers,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@router.get("/export/outbound")
def export_outbound(db: Session = Depends(get_db)):
    orders = db.query(OutboundOrder, Material.code, Material.description, Location.code.label('loc_code'), MaterialPriceVersion.batch_no.label('batch_no')).join(
        Material, OutboundOrder.material_id == Material.id
    ).join(
        Location, OutboundOrder.location_id == Location.id
    ).outerjoin(
        MaterialPriceVersion, OutboundOrder.price_version_id == MaterialPriceVersion.id
    ).all()
    
    data = []
    for order, mat_code, mat_desc, loc_code, batch_no in orders:
        data.append({
            "出库单号": order.order_no,
            "客户": order.customer,
            "领用人": order.receiver,
            "合同号": batch_no,
            "物料编码": mat_code,
            "物料描述": mat_desc,
            "库位编码": loc_code,
            "出库数量": order.quantity,
            "状态": order.status,
            "出库时间": order.outbound_time.strftime("%Y-%m-%d %H:%M:%S") if order.outbound_time else ""
        })
        
    df = pd.DataFrame(data)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='出库记录')
    
    output.seek(0)
    headers = {
        'Content-Disposition': 'attachment; filename="outbound_records.xlsx"'
    }
    return StreamingResponse(
        iter([output.getvalue()]),
        headers=headers,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
